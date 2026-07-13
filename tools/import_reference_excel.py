from __future__ import annotations

import argparse
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT))

from lab_accounting.core.config import load_config
from lab_accounting.core.fiscal_year import fiscal_year_for_date, target_month_for_date
from lab_accounting.repositories.database import Database


def main() -> int:
    parser = argparse.ArgumentParser(description="Reference Excel importer")
    parser.add_argument("excel_path", help="2026業務委託報酬.xlsx")
    args = parser.parse_args()

    excel_path = Path(args.excel_path)
    if not excel_path.exists():
        raise SystemExit(f"Excel file not found: {excel_path}")

    config = load_config()
    db = Database(config.database_path)
    db.initialize()

    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb["案件入力"]

    people: dict[str, int] = {}
    created_people = 0
    created_records = 0

    for row in ws.iter_rows(min_row=3, values_only=True):
        record_date, person_name, item_name, amount = row[:4]
        if not (record_date and person_name and item_name):
            continue
        if str(item_name).startswith("="):
            continue
        if person_name not in people:
            rows = db.query("SELECT person_id FROM persons WHERE name = ? LIMIT 1", (person_name,))
            if rows:
                people[str(person_name)] = int(rows[0]["person_id"])
            else:
                person_id = db.execute(
                    """
                    INSERT INTO persons (
                        name, display_name, color, resident_type, tax_rate,
                        daily_gross_amount, daily_net_amount, active
                    )
                    VALUES (?, ?, ?, 'resident', NULL, ?, ?, 1)
                    """,
                    (person_name, person_name, "#dbeafe", config.default_daily_gross_amount, config.default_daily_net_amount),
                )
                people[str(person_name)] = person_id
                created_people += 1

        if isinstance(record_date, datetime):
            day = record_date.date()
        elif hasattr(record_date, "date"):
            day = record_date.date()
        else:
            continue

        record_type = "carryover" if "残金" in str(item_name) or "残額" in str(item_name) else "income"
        db.execute(
            """
            INSERT INTO accounting_records (
                record_date, person_id, item_name, amount, record_type,
                fiscal_year, target_month, memo
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                day.isoformat(),
                people[str(person_name)],
                str(item_name),
                float(amount or 0),
                record_type,
                fiscal_year_for_date(day, config.fiscal_year_start_month),
                target_month_for_date(day),
                "Excel参照取込",
            ),
        )
        created_records += 1

    db.execute(
        """
        INSERT INTO operation_logs (event_level, event_type, message)
        VALUES ('INFO', 'REFERENCE_EXCEL_IMPORTED', ?)
        """,
        (f"Imported {created_records} records and {created_people} persons from {excel_path.name}",),
    )
    print(f"Imported persons: {created_people}")
    print(f"Imported records: {created_records}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
