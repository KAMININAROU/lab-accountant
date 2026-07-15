from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from lab_accounting.core.fiscal_year import fiscal_months, month_label
from lab_accounting.core.tax import calc_from_gross, effective_tax_rate


HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
NEGATIVE_FILL = PatternFill("solid", fgColor="F4CCCC")


class ExcelExporter:
    def __init__(self, resident_tax_rate: float, nonresident_tax_rate: float) -> None:
        self.resident_tax_rate = resident_tax_rate
        self.nonresident_tax_rate = nonresident_tax_rate

    def export(
        self,
        output_path: str,
        fiscal_year: int,
        records: list[dict],
        overview: list[dict],
        payments: list[dict],
        include_records: bool = True,
        include_summary: bool = True,
        include_payments: bool = True,
        include_items: bool = True,
        include_tax_sheet: bool = True,
        persons: list[dict] | None = None,
    ) -> str:
        wb = Workbook()
        wb.remove(wb.active)
        self._write_updated(wb)
        self._write_people(wb, persons or [])
        if include_records:
            self._write_records(wb, records)
        if include_summary:
            self._write_summary(wb, overview, fiscal_year)
        if include_payments:
            self._write_payment_details(wb, payments)
            self._write_payments(wb, payments, fiscal_year)
        if include_items:
            self._write_items(wb, records)
        if include_tax_sheet:
            self._write_tax(wb)

        path = Path(output_path)
        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(path)
        return str(path)

    def _write_updated(self, wb: Workbook) -> None:
        ws = wb.create_sheet("最終更新日")
        ws.append(["最終更新日"])
        ws.append([datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        ws.append([])
        ws.append(["更新者"])
        ws.append(["LabAccounting"])
        self._fit(ws)

    def _write_people(self, wb: Workbook, rows: list[dict]) -> None:
        ws = wb.create_sheet("人員管理")
        ws.append(
            [
                "人員ID",
                "氏名",
                "表示名",
                "カナ",
                "備考",
                "区分",
                "有効税率",
                "日額(支払総額)",
                "日額(自動手取)",
                "色",
                "状態",
            ]
        )
        for person in rows:
            rate = effective_tax_rate(person, self.resident_tax_rate, self.nonresident_tax_rate)
            daily_net = calc_from_gross(person["daily_gross_amount"], rate)["net_amount"]
            ws.append(
                [
                    person["person_id"],
                    person["name"],
                    person.get("display_name") or person["name"],
                    person.get("name_kana") or "",
                    person.get("memo") or "",
                    "非居住者" if person.get("resident_type") == "nonresident" else "居住者",
                    rate,
                    person["daily_gross_amount"],
                    daily_net,
                    person.get("color") or "",
                    "有効" if person.get("active") else "停止",
                ]
            )
        self._style_table(ws, money_cols=[8, 9])
        for cell in ws["G"][1:]:
            cell.number_format = "0.00%"

    def _write_records(self, wb: Workbook, rows: list[dict]) -> None:
        ws = wb.create_sheet("案件入力")
        headers = ["日付", "人", "名目", "金額", "種類", "年度", "対象月", "備考", "記録ID"]
        ws.append(headers)
        for row in rows:
            ws.append(
                [
                    row["record_date"],
                    row["person_name"],
                    row["item_name"],
                    row["amount"],
                    row["record_type"],
                    row["fiscal_year"],
                    row["target_month"],
                    row.get("memo") or "",
                    row["record_id"],
                ]
            )
        self._style_table(ws, money_cols=[4])

    def _write_summary(self, wb: Workbook, rows: list[dict], fiscal_year: int) -> None:
        ws = wb.create_sheet("Summary")
        ws.append([f"{fiscal_year}年度 集計"])
        ws.append(["人", "累積利用額", "累積支給日数", "累積支給額", "残額", "状態"])
        for row in rows:
            ws.append(
                [
                    row["person_name"],
                    row["total_added"],
                    row["payment_days"],
                    row["net_amount"],
                    row["balance"],
                    "有効" if row["active"] else "停止",
                ]
            )
            if row["balance"] < 0:
                for cell in ws[ws.max_row]:
                    cell.fill = NEGATIVE_FILL
        self._style_table(ws, header_row=2, money_cols=[2, 4, 5])
        ws["A1"].font = Font(bold=True, size=14)

    def _write_payment_details(self, wb: Workbook, rows: list[dict]) -> None:
        ws = wb.create_sheet("月次精算明細")
        ws.append(
            [
                "人",
                "年度",
                "対象月",
                "日数",
                "支払総額",
                "源泉徴収額",
                "手取額",
                "状態",
                "備考",
                "記録ID",
            ]
        )
        for row in rows:
            ws.append(
                [
                    row["person_name"],
                    row["fiscal_year"],
                    row["target_month"],
                    row["payment_days"],
                    row["gross_amount"],
                    row["withholding_amount"],
                    row["net_amount"],
                    row["status"],
                    row.get("memo") or "",
                    row["payment_id"],
                ]
            )
        self._style_table(ws, money_cols=[5, 6, 7])

    def _write_payments(self, wb: Workbook, rows: list[dict], fiscal_year: int) -> None:
        ws = wb.create_sheet("個人別支給額")
        months = fiscal_months(fiscal_year)
        by_person: dict[int, dict] = {}
        for row in rows:
            person_id = int(row["person_id"])
            person = by_person.setdefault(person_id, {"name": row["person_name"], "months": {}})
            month = person["months"].setdefault(row["target_month"], {"payment_days": 0.0, "net_amount": 0.0})
            month["payment_days"] += float(row.get("payment_days") or 0)
            month["net_amount"] += float(row.get("net_amount") or 0)

        ordered_people = sorted(by_person.items(), key=lambda item: (str(item[1]["name"]).casefold(), item[0]))
        ws.append([f"{fiscal_year}年度 業務委託報酬"])
        ws.append(["日数"])
        ws.append(["人"] + [month_label(month) for month in months] + ["合計"])
        for _, person in ordered_people:
            values = [float(person["months"].get(month, {}).get("payment_days", 0) or 0) for month in months]
            ws.append([person["name"]] + values + [sum(values)])
        ws.append([])
        ws.append(["金額"])
        amount_header_row = ws.max_row + 1
        ws.append(["人"] + [month_label(month) for month in months] + ["合計"])
        for _, person in ordered_people:
            values = [float(person["months"].get(month, {}).get("net_amount", 0) or 0) for month in months]
            ws.append([person["name"]] + values + [sum(values)])
        self._style_headers(ws, [3, amount_header_row])
        for row in ws.iter_rows(min_row=amount_header_row + 1, min_col=2):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0"
        self._fit(ws)
        ws.freeze_panes = "B4"

    def _write_items(self, wb: Workbook, rows: list[dict]) -> None:
        ws = wb.create_sheet("使用費目一覧")
        ws.append(["年度", "対象月", "人", "日付", "名目", "金額", "種類", "備考", "記録ID"])
        for row in rows:
            ws.append(
                [
                    row["fiscal_year"],
                    row["target_month"],
                    row["person_name"],
                    row["record_date"],
                    row["item_name"],
                    row["amount"],
                    row["record_type"],
                    row.get("memo") or "",
                    row["record_id"],
                ]
            )
        self._style_table(ws, money_cols=[6])

    def _write_tax(self, wb: Workbook) -> None:
        ws = wb.create_sheet("税額計算表")
        ws.append(["税額計算参考"])
        ws.append([])
        ws.append(["区分", "税率", "支払総額", "源泉徴収額", "手取額"])
        for label, rate, gross in [
            ("居住者", self.resident_tax_rate, 10000),
            ("非居住者", self.nonresident_tax_rate, 10000),
        ]:
            result = calc_from_gross(gross, rate)
            ws.append([label, rate, result["gross_amount"], result["withholding_amount"], result["net_amount"]])
        self._style_table(ws, header_row=3, money_cols=[3, 4, 5])
        for cell in ws["B"]:
            if isinstance(cell.value, float):
                cell.number_format = "0.00%"

    def _style_table(self, ws, header_row: int = 1, money_cols: list[int] | None = None) -> None:
        self._style_headers(ws, [header_row])
        for col in money_cols or []:
            for row in range(header_row + 1, ws.max_row + 1):
                ws.cell(row, col).number_format = "#,##0"
        ws.freeze_panes = f"A{header_row + 1}"
        ws.auto_filter.ref = ws.dimensions
        self._fit(ws)

    @staticmethod
    def _style_headers(ws, rows: list[int]) -> None:
        for row in rows:
            for cell in ws[row]:
                cell.font = Font(bold=True)
                cell.fill = HEADER_FILL

    @staticmethod
    def _fit(ws) -> None:
        for column in ws.columns:
            max_len = 10
            col_letter = get_column_letter(column[0].column)
            for cell in column:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, min(len(value) + 2, 36))
            ws.column_dimensions[col_letter].width = max_len
